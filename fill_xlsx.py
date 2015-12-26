import openpyxl

wb = openpyxl.load_workbook('analysis_template.xlsx')
ws = wb.get_sheet_by_name("template")

notes = open('notes.txt').read().split('\n')[:-1]

ws.cell('W15').value = notes[2] # k_prior

# t_pi_prior
ws.cell('W10').value = notes[4] 
ws.cell('W11').value = notes[5]
ws.cell('W12').value = notes[6]

ws.cell('B1').value = notes[10] # delta_1
ws.cell('C1').value = notes[12] # delta_2
ws.cell('D1').value = notes[14] # k
ws.cell('E1').value = notes[16] # lambda
ws.cell('F1').value = notes[18] # nu

# t_pi
ws.cell('N1').value = notes[21]
ws.cell('O1').value = notes[22]
ws.cell('P1').value = notes[23] 

ws.cell('W3').value = notes[26] # G
ws.cell('W4').value = notes[28] # n
ws.cell('W5').value = notes[30] # times
 
# b
ws.cell('W17').value = notes[33]
ws.cell('W18').value = notes[34]
ws.cell('W19').value = notes[35]

psi = open('psi.txt').read().split('\n')[:-1]
t_pi = open('t_pi.txt').read().split('\n')[:-1]
accuracy = open('accuracy.txt').read().split('\n')[:-1]
adjusted_rand = open('adjusted_rand.txt').read().split('\n')[:-1]

ws.cell('A9').value = psi[0]
ws.cell('N9').value = t_pi[0]
ws.cell('AA8').value = accuracy[0]
ws.cell('AO8').value = adjusted_rand[0]

for i in range(100):
	row_data = psi[i+1].split('\t')
	for j in range(6):
		ws.cell(row = 9+i, column = j).value = row_data[j]
	row_data = t_pi[i+1].split('\t')
	for j in range(3):
		ws.cell(row = 9+i, column = 13+j).value = row_data[j]
	row_data = accuracy[i+1].split('\t')
	for j in range(9):
		ws.cell(row = 8+i, column = 26+j).value = row_data[j]
	row_data = adjusted_rand[i+1].split('\t')
	for j in range(5):
		ws.cell(row = 8+i, column = 40+j).value = row_data[j]

i = 100
row_data = accuracy[i+1].split('\t')
for j in range(9):
	ws.cell(row = 8+i, column = 26+j).value = row_data[j]
row_data = adjusted_rand[i+1].split('\t')
for j in range(5):
	ws.cell(row = 8+i, column = 40+j).value = row_data[j]

wb.save('analysis_template.xlsx')
