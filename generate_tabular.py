import openpyxl

wb = openpyxl.load_workbook('analysis_template.xlsx', data_only=True)
ws = wb.get_sheet_by_name("template")

psi_tabular = ['']*9

psi_tabular[0] = ''' \\begin{tabular}{c c c c c c}'''
psi_tabular[1] = ''' & $\\delta_1$ & $\\delta_2$ & $k$ & $\\lambda$ & $\\nu$ \\\\ '''

psi_tabular[2] = ''' true value '''
for i in range(5):
	psi_tabular[2] += '& ' + str("{:.4E}".format(ws.cell(row = 0, column = 1 + i).value)) + ' '
psi_tabular[2] += ''' \\\\ '''

psi_tabular[3] = ''' average '''
psi_tabular[4] = ''' median '''
psi_tabular[5] = ''' mse '''
psi_tabular[6] = ''' average bias '''
psi_tabular[7] = ''' median bias '''

for j in range(5):
	row = 3+j
	for i in range(5):
		psi_tabular[row] += '& ' + str("{:.4E}".format(ws.cell(row = 2+j, column = 1+i).value))+' '
	psi_tabular[row] += ''' \\\\ '''

psi_tabular[8] = ''' \\end{tabular} '''





t_pi_tabular = ['']*9

t_pi_tabular[0] = ''' \\begin{tabular}{c c c c}'''
t_pi_tabular[1] = ''' & $\\pi_1$ & $\\pi_2$ & $\\pi_3$ \\\\ '''

t_pi_tabular[2] = ''' true value '''
for i in range(3):
	t_pi_tabular[2] += '& ' + str("{:.4E}".format(ws.cell(row = 0, column = 13 + i).value)) + ' '
t_pi_tabular[2] += ''' \\\\ '''

t_pi_tabular[3] = ''' average '''
t_pi_tabular[4] = ''' median '''
t_pi_tabular[5] = ''' mse '''
t_pi_tabular[6] = ''' average bias '''
t_pi_tabular[7] = ''' median bias '''

for j in range(5):
	row = 3+j
	for i in range(3):
		t_pi_tabular[row] += '& ' + str("{:.4E}".format(ws.cell(row = 2+j, column = 13+i).value))+' '
	t_pi_tabular[row] += ''' \\\\ '''

t_pi_tabular[8] = ''' \\end{tabular} '''





fpr_fnr_fdr_fndr_tabular = ['']*5

fpr_fnr_fdr_fndr_tabular[0] = ''' \\begin{tabular}{c c c c c} '''
fpr_fnr_fdr_fndr_tabular[1] = ''' & FPR & FNR & FDR & FNDR \\\\ '''

fpr_fnr_fdr_fndr_tabular[2] = ''' average '''
fpr_fnr_fdr_fndr_tabular[3] = ''' median '''

for j in range(2):
	row = 2+j
	for i in range(4):
		fpr_fnr_fdr_fndr_tabular[row] += '& ' + str("{:.4E}".format(ws.cell(row = 1+j, column = 35+i).value))+' '
	fpr_fnr_fdr_fndr_tabular[row] += ''' \\\\ '''

fpr_fnr_fdr_fndr_tabular[4] = ''' \\end{tabular} '''

adjusted_rand_tabuler = [''] * 5
adjusted_rand_tabuler[0] = ''' \\begin{tabular}{c c c c c c} '''
adjusted_rand_tabuler[1] = ''' & Rand & HA & MA & FM & Jaccard \\\\ '''
adjusted_rand_tabuler[2] = ''' average '''
adjusted_rand_tabuler[3] = ''' median '''

for j in range(2):
	row = 2+j
	for i in range(5):
		adjusted_rand_tabuler[row] += '& ' + str("{:.4E}".format(ws.cell(row = 1+j, column = 40+i).value))+' '
	adjusted_rand_tabuler[row] += ''' \\\\ '''

adjusted_rand_tabuler[4] = ''' \\end{tabular} '''

fout = open('tabular.txt',"w")
fout.write('\n'.join(psi_tabular))
fout.write('\n\n')
fout.write('\n'.join(t_pi_tabular))
fout.write('\n\n')
fout.write('\n'.join(fpr_fnr_fdr_fndr_tabular))
fout.write('\n\n')
fout.write('\n'.join(adjusted_rand_tabuler))
fout.write('\n\n')
fout.close()
